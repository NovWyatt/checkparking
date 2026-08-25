"""Read a Check Vehicle OCR export back into the Results workspace.

The importer is intentionally limited to the app's own export sheets.  It
never modifies the workbook and does not treat arbitrary Excel files as OCR
results.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import ImageResult, PlateCandidate


class ResultsImportError(RuntimeError):
    """Raised when a workbook is not a supported Check Vehicle OCR export."""


def load_exported_results(workbook_path: Path) -> list[ImageResult]:
    """Return result rows from a non-destructively opened OCR export workbook."""

    path = Path(workbook_path).expanduser()
    if not path.is_file():
        raise ResultsImportError("Không tìm thấy file Excel đã chọn.")
    if path.suffix.lower() != ".xlsx":
        raise ResultsImportError("Chỉ hỗ trợ file Excel .xlsx do Check Vehicle OCR đã xuất.")

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ResultsImportError("Không thể đọc file Excel. Hãy chọn file do Check Vehicle OCR đã xuất.") from exc
    try:
        worksheet = workbook["Theo_tung_anh"] if "Theo_tung_anh" in workbook.sheetnames else None
        if worksheet is None:
            raise ResultsImportError("File này không có sheet Theo_tung_anh của Check Vehicle OCR.")
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ResultsImportError("Sheet Theo_tung_anh chưa có dữ liệu để mở lại.")
        columns = _columns(headers)
        required = {"Tên file", "Số biển số", "Trạng thái", "Đường dẫn ảnh"}
        missing = required - set(columns)
        if missing:
            raise ResultsImportError("Sheet Theo_tung_anh không đúng cấu trúc file xuất của Check Vehicle OCR.")
        plate_columns = [(name, index) for name, index in columns.items() if re.fullmatch(r"Biển số \d+", name)]
        if not plate_columns:
            raise ResultsImportError("Sheet Theo_tung_anh không có cột Biển số để khôi phục kết quả.")
        crop_paths = _read_crop_paths(workbook)

        imported: list[ImageResult] = []
        for row in rows:
            values = list(row)
            file_name = _value(values, columns["Tên file"])
            image_value = _value(values, columns["Đường dẫn ảnh"])
            if not file_name and not image_value:
                continue
            image_path = Path(image_value).expanduser() if image_value else path.parent / file_name
            plates = [_plate(_value(values, index)) for _name, index in plate_columns]
            plates = [plate for plate in plates if plate is not None]
            for plate in plates:
                available_crops = crop_paths.get(_crop_key(image_path, plate.final_text), [])
                while available_crops:
                    crop_path = available_crops.pop(0)
                    if crop_path.is_file():
                        plate.crop_path = crop_path
                        break
            width, height = _size(_value(values, columns.get("Kích thước", -1)))
            imported.append(
                ImageResult(
                    image_path=image_path,
                    status=_value(values, columns["Trạng thái"]) or "IMPORTED",
                    reason=_value(values, columns.get("Lý do/Ghi chú", -1)) or "Đã mở lại từ file Excel đã xuất.",
                    blur_score=_number(_value(values, columns.get("Độ nét ảnh", -1))),
                    width=width,
                    height=height,
                    candidate_count=len(plates),
                    plates=plates,
                )
            )
    finally:
        workbook.close()

    if not imported:
        raise ResultsImportError("File Excel không có dòng ảnh nào để khôi phục vào Kết quả.")
    return imported


def _read_crop_paths(workbook) -> dict[tuple[str, str], list[Path]]:
    """Map export rows to their existing crop files without changing Excel."""

    if "Bien_so_doc_duoc" not in workbook.sheetnames:
        return {}
    worksheet = workbook["Bien_so_doc_duoc"]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return {}
    columns = _columns(headers)
    required = {"Đường dẫn ảnh", "Biển số xuất Excel", "Crop biển số"}
    if not required.issubset(columns):
        return {}

    crop_paths: dict[tuple[str, str], list[Path]] = {}
    for row in rows:
        values = list(row)
        image_value = _value(values, columns["Đường dẫn ảnh"])
        plate_value = _value(values, columns["Biển số xuất Excel"])
        crop_value = _value(values, columns["Crop biển số"])
        if not image_value or not plate_value or not crop_value:
            continue
        key = _crop_key(Path(image_value).expanduser(), plate_value)
        crop_paths.setdefault(key, []).append(Path(crop_value).expanduser())
    return crop_paths


def _crop_key(image_path: Path, plate_text: str) -> tuple[str, str]:
    """Return a stable Windows-friendly key for mapping an exported plate."""

    image_key = str(image_path).strip().replace("/", "\\").casefold()
    plate_key = re.sub(r"\s+", "", plate_text).casefold()
    return image_key, plate_key


def _columns(headers: tuple[object, ...]) -> dict[str, int]:
    return {str(value).strip(): index for index, value in enumerate(headers) if str(value or "").strip()}


def _value(values: list[object], index: int) -> str:
    if index < 0 or index >= len(values) or values[index] is None:
        return ""
    return str(values[index]).strip()


def _number(value: str) -> float:
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return 0.0


def _size(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"\s*(\d+)\s*x\s*(\d+)\s*", value, flags=re.IGNORECASE)
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def _plate(value: str) -> PlateCandidate | None:
    if not value:
        return None
    return PlateCandidate(
        bbox=(0, 0, 0, 0),
        score=0.0,
        source="excel_import",
        text=value,
        raw_text=value,
        readable=True,
        reason="Khôi phục từ file Excel đã xuất.",
    )
