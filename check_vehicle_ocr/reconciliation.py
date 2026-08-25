"""Local Excel reconciliation for reviewed vehicle-plate exports.

The module deliberately keeps comparison rules visible in the generated
workbook.  It never edits either source workbook and only accepts a near
match when there is one unique candidate with a single-character difference.
"""

from __future__ import annotations

import os
import unicodedata
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
SUCCESS_FILL = PatternFill("solid", fgColor="E2F0D9")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")


class ReconciliationError(RuntimeError):
    """Raised when a workbook cannot safely be used for reconciliation."""


@dataclass(frozen=True)
class PlateRecord:
    source: str
    workbook_path: Path
    sheet_name: str
    row_number: int
    raw_text: str
    normalized_text: str

    @property
    def location(self) -> str:
        return f"{self.sheet_name}!{self.row_number}"


@dataclass(frozen=True)
class SourceRecords:
    source: str
    workbook_path: Path
    sheet_name: str
    plate_column: str
    records: tuple[PlateRecord, ...]


@dataclass(frozen=True)
class PlateMatch:
    status: str
    target: PlateRecord | None = None
    note: str = ""
    candidates: tuple[PlateRecord, ...] = ()

    @property
    def is_accepted(self) -> bool:
        return self.status in {"exact", "near"}

    @property
    def needs_review(self) -> bool:
        return self.status in {"ambiguous", "review"}


@dataclass(frozen=True)
class ReconciliationRow:
    ocr: PlateRecord
    fee: PlateMatch
    software: PlateMatch
    conclusion: str


@dataclass(frozen=True)
class ReconciliationReport:
    ocr_source: SourceRecords
    fee_source: SourceRecords
    software_source: SourceRecords | None
    suffix_length: int
    rows: tuple[ReconciliationRow, ...]
    fee_surplus: tuple[PlateRecord, ...]
    software_surplus: tuple[PlateRecord, ...]
    duplicates: tuple[tuple[str, str, tuple[PlateRecord, ...]], ...]

    @property
    def compare_software(self) -> bool:
        return self.software_source is not None


def normalize_plate(value: object) -> str:
    """Return a stable comparison key without changing source data."""

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text or text.startswith("="):
        return ""
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(character for character in decomposed.upper() if character.isascii() and character.isalnum())


def write_reconciliation_template(output_path: Path, source_label: str) -> Path:
    """Create a paste-ready source file with a single unambiguous plate column."""

    workbook = Workbook()
    data = workbook.active
    data.title = "Danh_sach"
    data.append(["Biển số", "Ghi chú"])
    data.freeze_panes = "A2"
    data.auto_filter.ref = "A1:B1"
    data.column_dimensions["A"].width = 22
    data.column_dimensions["B"].width = 44

    guide = workbook.create_sheet("Huong_dan")
    guide.append(["Hướng dẫn"])
    guide.append([f"Dán biển số {source_label.lower()} vào cột A của sheet Danh_sach, bắt đầu từ dòng 2."])
    guide.append(["Không gộp ô trong vùng dữ liệu. Cột Ghi chú là tùy chọn và không được dùng để đối chiếu."])
    guide.column_dimensions["A"].width = 100

    _style_workbook(workbook)
    return _atomic_save(workbook, output_path)


def read_plate_records(workbook_path: Path, source: str) -> SourceRecords:
    """Read the most suitable plate column from an .xlsx workbook.

    The function recognizes the application's ``Biển số xuất Excel`` column
    and the ``Biển số`` column in the provided templates.  It purposefully
    fails when no clear plate column is present rather than guessing a number
    column such as "Số biển số".
    """

    path = Path(workbook_path).expanduser()
    if not path.is_file():
        raise ReconciliationError(f"Không tìm thấy file {source.lower()}: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ReconciliationError(f"File {source.lower()} phải là Excel .xlsx.")

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ReconciliationError(f"Không mở được file {source.lower()}. Hãy kiểm tra file Excel có bị lỗi hoặc đang mã hóa hay không.") from exc

    try:
        candidate: tuple[int, int, int, object] | None = None
        for worksheet in workbook.worksheets:
            for row_number, values in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
                for column_number, value in enumerate(values, start=1):
                    score = _plate_header_score(value)
                    if score <= 0:
                        continue
                    current = (score, row_number, column_number, worksheet)
                    if candidate is None or current[:3] > candidate[:3]:
                        candidate = current

        if candidate is None:
            raise ReconciliationError(
                f"Không tìm thấy cột biển số trong file {source.lower()}. "
                "Hãy dùng file mẫu hoặc đặt tiêu đề cột là 'Biển số'."
            )

        _score, header_row, plate_column, worksheet = candidate
        header_value = worksheet.cell(header_row, plate_column).value
        records: list[PlateRecord] = []
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, min_col=plate_column, max_col=plate_column, values_only=True),
            start=header_row + 1,
        ):
            raw_value = values[0] if values else None
            raw_text = _cell_text(raw_value)
            normalized_text = normalize_plate(raw_text)
            if not normalized_text:
                continue
            records.append(
                PlateRecord(
                    source=source,
                    workbook_path=path.resolve(),
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    raw_text=raw_text,
                    normalized_text=normalized_text,
                )
            )
    finally:
        workbook.close()

    if not records:
        raise ReconciliationError(f"Cột biển số trong file {source.lower()} chưa có dữ liệu để đối chiếu.")
    return SourceRecords(
        source=source,
        workbook_path=path.resolve(),
        sheet_name=worksheet.title,
        plate_column=str(header_value or "Biển số"),
        records=tuple(records),
    )


def reconcile_workbooks(
    ocr_path: Path,
    fee_path: Path,
    software_path: Path | None = None,
    *,
    suffix_length: int = 4,
) -> ReconciliationReport:
    """Apply the operator workflow: fee first, then optional software."""

    if suffix_length not in {3, 4}:
        raise ValueError("suffix_length chỉ hỗ trợ 3 hoặc 4.")

    ocr_source = read_plate_records(ocr_path, "OCR")
    fee_source = read_plate_records(fee_path, "Báo phí")
    software_source = read_plate_records(software_path, "Phần mềm") if software_path else None
    fee_index = _build_index(fee_source.records)
    software_index = _build_index(software_source.records) if software_source else None

    rows: list[ReconciliationRow] = []
    for ocr in ocr_source.records:
        fee_match = _find_match(ocr, fee_index, suffix_length)
        if software_index is None:
            software_match = PlateMatch("skipped", note="Không chọn đối chiếu phần mềm.")
        elif fee_match.is_accepted:
            software_match = PlateMatch("skipped", note="Không dò vì đã khớp báo phí.")
        else:
            software_match = _find_match(ocr, software_index, suffix_length)
        rows.append(ReconciliationRow(ocr=ocr, fee=fee_match, software=software_match, conclusion=_conclusion(fee_match, software_match)))

    accepted_fee = {row.fee.target.normalized_text for row in rows if row.fee.is_accepted and row.fee.target}
    accepted_software = {row.software.target.normalized_text for row in rows if row.software.is_accepted and row.software.target}
    duplicates = tuple(
        _duplicate_groups(ocr_source.records)
        + _duplicate_groups(fee_source.records)
        + (_duplicate_groups(software_source.records) if software_source else [])
    )
    return ReconciliationReport(
        ocr_source=ocr_source,
        fee_source=fee_source,
        software_source=software_source,
        suffix_length=suffix_length,
        rows=tuple(rows),
        fee_surplus=tuple(record for record in fee_source.records if record.normalized_text not in accepted_fee),
        software_surplus=tuple(record for record in software_source.records if record.normalized_text not in accepted_software) if software_source else (),
        duplicates=duplicates,
    )


def export_reconciliation_report(report: ReconciliationReport, output_path: Path) -> Path:
    """Write a non-destructive, reviewable reconciliation workbook."""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Tổng_quan"
    _write_summary(summary, report)

    all_rows = workbook.create_sheet("Kết_quả_chính")
    _write_result_rows(all_rows, report.rows)

    fee_exact = workbook.create_sheet("Khớp_báo_phí")
    _write_result_rows(fee_exact, [row for row in report.rows if row.fee.status == "exact"])

    fee_near = workbook.create_sheet("Khớp_gần_báo_phí")
    _write_result_rows(fee_near, [row for row in report.rows if row.fee.status == "near"])

    review = workbook.create_sheet("Cần_xác_nhận")
    _write_result_rows(
        review,
        [row for row in report.rows if row.fee.needs_review or row.software.needs_review],
    )

    if report.compare_software:
        software_only = workbook.create_sheet("Phần_mềm_không_báo_phí")
        _write_result_rows(
            software_only,
            [row for row in report.rows if row.fee.status == "missing" and row.software.is_accepted],
        )
        neither = workbook.create_sheet("Không_có_cả_hai")
        _write_result_rows(
            neither,
            [row for row in report.rows if row.fee.status == "missing" and row.software.status == "missing"],
        )

    duplicates = workbook.create_sheet("Trùng_lặp")
    _write_duplicates(duplicates, report.duplicates)

    fee_surplus = workbook.create_sheet("Dư_báo_phí")
    _write_source_records(fee_surplus, report.fee_surplus)
    if report.compare_software:
        software_surplus = workbook.create_sheet("Dư_phần_mềm")
        _write_source_records(software_surplus, report.software_surplus)

    _style_workbook(workbook)
    return _atomic_save(workbook, output_path)


def _plate_header_score(value: object) -> int:
    header = normalize_plate(value)
    if not header or "SOBIENSO" in header:
        return 0
    if header == "BIENSOXUATEXCEL":
        return 120
    if header == "BIENSO":
        return 110
    if header in {"BIENSOOCRCHUANHOA", "BIENSOOCR"}:
        return 100
    if "BIENSO" in header:
        return 80
    return 0


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _build_index(records: tuple[PlateRecord, ...]) -> tuple[dict[str, list[PlateRecord]], dict[str, list[PlateRecord]]]:
    exact: dict[str, list[PlateRecord]] = defaultdict(list)
    deletions: dict[str, list[PlateRecord]] = defaultdict(list)
    for record in records:
        exact[record.normalized_text].append(record)
        for signature in _deletion_signatures(record.normalized_text):
            deletions[signature].append(record)
    return exact, deletions


def _deletion_signatures(value: str) -> set[str]:
    return {value[:index] + value[index + 1 :] for index in range(len(value))}


def _find_match(
    record: PlateRecord,
    index: tuple[dict[str, list[PlateRecord]], dict[str, list[PlateRecord]]],
    suffix_length: int,
) -> PlateMatch:
    exact, deletions = index
    exact_matches = exact.get(record.normalized_text, [])
    if exact_matches:
        target = exact_matches[0]
        note = "Khớp hoàn toàn sau khi bỏ khoảng trắng và ký tự định dạng."
        if len(exact_matches) > 1:
            note += f" Nguồn đối chiếu có {len(exact_matches)} dòng trùng biển này."
        return PlateMatch("exact", target=target, note=note, candidates=tuple(exact_matches))

    candidates: dict[tuple[str, str, int], PlateRecord] = {}
    for signature in _deletion_signatures(record.normalized_text):
        for target in exact.get(signature, []):
            candidates[(str(target.workbook_path), target.sheet_name, target.row_number)] = target
        for target in deletions.get(signature, []):
            candidates[(str(target.workbook_path), target.sheet_name, target.row_number)] = target
    for target in deletions.get(record.normalized_text, []):
        candidates[(str(target.workbook_path), target.sheet_name, target.row_number)] = target

    one_edit = [target for target in candidates.values() if _levenshtein_distance(record.normalized_text, target.normalized_text) == 1]
    if not one_edit:
        return PlateMatch("missing", note="Không tìm thấy biển số trong nguồn này.")
    if len(one_edit) > 1:
        return PlateMatch(
            "ambiguous",
            note=f"Có {len(one_edit)} biển gần đúng, không tự chấp nhận để tránh ghép nhầm.",
            candidates=tuple(sorted(one_edit, key=lambda item: (item.sheet_name, item.row_number))),
        )

    target = one_edit[0]
    tail_matches = _tail_alignment_matches(record.normalized_text, target.normalized_text, suffix_length)
    difference = _difference_note(record.normalized_text, target.normalized_text)
    if tail_matches >= suffix_length - 1:
        return PlateMatch(
            "near",
            target=target,
            note=(
                f"Khớp gần được chấp nhận: {difference}; trùng {tail_matches}/{suffix_length} ký tự cuối "
                "sau khi canh chỉnh."
            ),
            candidates=(target,),
        )
    return PlateMatch(
        "review",
        target=target,
        note=(
            f"{difference}; chỉ trùng {tail_matches}/{suffix_length} ký tự cuối sau khi canh chỉnh "
            "nên cần xác nhận thủ công."
        ),
        candidates=(target,),
    )


def _levenshtein_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    if abs(len(left) - len(right)) > 1:
        return max(len(left), len(right))
    previous = list(range(len(right) + 1))
    for left_index, left_character in enumerate(left, start=1):
        current = [left_index]
        for right_index, right_character in enumerate(right, start=1):
            current.append(min(current[-1] + 1, previous[right_index] + 1, previous[right_index - 1] + (left_character != right_character)))
        previous = current
    return previous[-1]


def _tail_alignment_matches(left: str, right: str, suffix_length: int) -> int:
    """Return the shared tail length while tolerating one inserted/deleted character.

    OCR is the reviewed source of truth.  A one-character omission or addition
    in a fee/software workbook shifts tail positions, so positional comparison
    would reject a valid unique match.  The longest common subsequence keeps the
    same conservative ``suffix_length - 1`` threshold without that false reject.
    """

    left_tail = left[-suffix_length:]
    right_tail = right[-suffix_length:]
    previous = [0] * (len(right_tail) + 1)
    for left_character in left_tail:
        current = [0]
        for index, right_character in enumerate(right_tail, start=1):
            if left_character == right_character:
                current.append(previous[index - 1] + 1)
            else:
                current.append(max(previous[index], current[-1]))
        previous = current
    return previous[-1]


def _difference_note(left: str, right: str) -> str:
    if len(left) != len(right):
        return "Thiếu hoặc dư 1 ký tự"
    for index, (first, second) in enumerate(zip(left, right), start=1):
        if first != second:
            return f"Khác 1 ký tự ở vị trí {index}: {first}/{second}"
    return "Khác 1 ký tự"


def _conclusion(fee: PlateMatch, software: PlateMatch) -> str:
    if fee.status == "exact":
        return "Khớp báo phí"
    if fee.status == "near":
        return "Khớp gần báo phí"
    if fee.needs_review:
        return "Cần xác nhận báo phí"
    if software.status == "exact":
        return "Có phần mềm, không báo phí"
    if software.status == "near":
        return "Có gần đúng phần mềm, không báo phí"
    if software.needs_review:
        return "Cần xác nhận phần mềm"
    if software.status == "skipped":
        return "Không có báo phí"
    return "Không có trên cả hai nguồn"


def _duplicate_groups(records: tuple[PlateRecord, ...]) -> list[tuple[str, str, tuple[PlateRecord, ...]]]:
    grouped: dict[str, list[PlateRecord]] = defaultdict(list)
    for record in records:
        grouped[record.normalized_text].append(record)
    return [
        (records_for_plate[0].source, normalized, tuple(records_for_plate))
        for normalized, records_for_plate in grouped.items()
        if len(records_for_plate) > 1
    ]


def _write_summary(sheet, report: ReconciliationReport) -> None:
    rows = report.rows
    values = [
        ("Thời gian xuất", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("File OCR", str(report.ocr_source.workbook_path)),
        ("File báo phí", str(report.fee_source.workbook_path)),
        ("File phần mềm", str(report.software_source.workbook_path) if report.software_source else "Không đối chiếu"),
        ("Quy tắc đuôi", f"Đối chiếu gần theo {report.suffix_length} ký tự cuối"),
        ("Tổng biển OCR", len(rows)),
        ("Khớp báo phí", sum(row.fee.status == "exact" for row in rows)),
        ("Khớp gần báo phí", sum(row.fee.status == "near" for row in rows)),
        ("Có phần mềm, không báo phí", sum(row.fee.status == "missing" and row.software.status == "exact" for row in rows)),
        ("Có gần đúng phần mềm, không báo phí", sum(row.fee.status == "missing" and row.software.status == "near" for row in rows)),
        ("Không có trên cả hai nguồn", sum(row.fee.status == "missing" and row.software.status == "missing" for row in rows)),
        ("Cần xác nhận", sum(row.fee.needs_review or row.software.needs_review for row in rows)),
        ("Nhóm biển trùng", len(report.duplicates)),
    ]
    sheet.append(["Chỉ mục", "Giá trị"])
    for row in values:
        sheet.append(row)


def _write_result_rows(sheet, rows: list[ReconciliationRow] | tuple[ReconciliationRow, ...]) -> None:
    sheet.append(
        [
            "STT",
            "Biển số OCR",
            "Chuẩn hóa OCR",
            "Dòng OCR",
            "Kết quả báo phí",
            "Biển số báo phí",
            "Dòng báo phí",
            "Ghi chú báo phí",
            "Kết quả phần mềm",
            "Biển số phần mềm",
            "Dòng phần mềm",
            "Ghi chú phần mềm",
            "Kết luận",
        ]
    )
    for index, row in enumerate(rows, start=1):
        sheet.append(
            [
                index,
                row.ocr.raw_text,
                row.ocr.normalized_text,
                row.ocr.location,
                _match_label(row.fee.status),
                row.fee.target.raw_text if row.fee.target else "",
                row.fee.target.location if row.fee.target else "",
                row.fee.note,
                _match_label(row.software.status),
                row.software.target.raw_text if row.software.target else "",
                row.software.target.location if row.software.target else "",
                row.software.note,
                row.conclusion,
            ]
        )


def _write_duplicates(sheet, groups: tuple[tuple[str, str, tuple[PlateRecord, ...]], ...]) -> None:
    sheet.append(["Nguồn", "Biển số chuẩn hóa", "Số dòng", "Biển số gốc", "Vị trí dòng"])
    for source, normalized, records in groups:
        sheet.append(
            [
                source,
                normalized,
                len(records),
                "; ".join(record.raw_text for record in records),
                "; ".join(record.location for record in records),
            ]
        )


def _write_source_records(sheet, records: tuple[PlateRecord, ...]) -> None:
    sheet.append(["STT", "Biển số", "Chuẩn hóa", "Sheet", "Dòng"])
    for index, record in enumerate(records, start=1):
        sheet.append([index, record.raw_text, record.normalized_text, record.sheet_name, record.row_number])


def _match_label(status: str) -> str:
    labels = {
        "exact": "Khớp hoàn toàn",
        "near": "Khớp gần",
        "missing": "Không có",
        "ambiguous": "Nhiều ứng viên",
        "review": "Cần xác nhận",
        "skipped": "Không dò",
    }
    return labels.get(status, status)


def _style_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            _escape_formula(cell)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            text_values = {str(cell.value or "") for cell in row}
            fill = None
            if any("Cần xác nhận" in value or "Nhiều ứng viên" in value for value in text_values):
                fill = ERROR_FILL
            elif any("Khớp gần" in value for value in text_values):
                fill = WARNING_FILL
            elif any("Khớp hoàn toàn" in value or value == "Khớp báo phí" for value in text_values):
                fill = SUCCESS_FILL
            for cell in row:
                _escape_formula(cell)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    cell.fill = fill
        for column in sheet.columns:
            length = max((len(str(cell.value or "")) for cell in column), default=10)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(length + 2, 12), 58)


def _escape_formula(cell) -> None:
    value = cell.value
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        cell.value = "'" + value


def _atomic_save(workbook: Workbook, output_path: Path) -> Path:
    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.parent / f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    except PermissionError as exc:
        raise ReconciliationError("Không thể ghi file Excel. Hãy đóng file đang mở trong Microsoft Excel rồi thử lại.") from exc
    finally:
        if temporary.exists():
            temporary.unlink(missing_ok=True)
        workbook.close()
    return path
