from __future__ import annotations

import hashlib
import os
import tempfile
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage, ImageOps

from .models import ImageResult, PlateCandidate
from .plate_formatting import DetectedPlateFormat, PlateFormatStatus, PlateType, coerce_plate_type

try:
    from pillow_heif import register_heif_opener

    register_heif_opener()
except Exception:
    pass


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")


class ExcelExportError(RuntimeError):
    pass


def export_results(
    results: list[ImageResult],
    output_path: Path,
    blur_threshold: float,
    reviewed: bool = False,
    include_images: bool = True,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="check_vehicle_excel_") as temp_dir:
        workbook = Workbook()
        media_cache: dict[tuple[str, int, int], Path] | None = {} if include_images else None

        summary = workbook.active
        summary.title = "Tong_quan"
        _write_summary(summary, results, blur_threshold, reviewed)

        per_image_sheet = workbook.create_sheet("Theo_tung_anh")
        _write_per_image(per_image_sheet, results, media_cache, Path(temp_dir))

        readable_sheet = workbook.create_sheet("Bien_so_doc_duoc")
        _write_readable(readable_sheet, results, reviewed, media_cache, Path(temp_dir))

        special_sheet = workbook.create_sheet("Bien_so_dac_biet")
        _write_special_plates(special_sheet, results)

        review_sheet = workbook.create_sheet("Can_kiem_tra")
        _write_review(review_sheet, results, blur_threshold, reviewed, media_cache, Path(temp_dir))

        if reviewed:
            review_all_sheet = workbook.create_sheet("Review_tat_ca")
            _write_review_all(review_all_sheet, results, media_cache, Path(temp_dir))

        all_sheet = workbook.create_sheet("Tat_ca_anh")
        _write_all_images(all_sheet, results, media_cache, Path(temp_dir))

        for sheet in workbook.worksheets:
            _style_sheet(sheet)

        temporary_path = output_path.parent / f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
        try:
            workbook.save(temporary_path)
            os.replace(temporary_path, output_path)
        except PermissionError as exc:
            raise ExcelExportError("Không thể ghi file Excel. Hãy đóng file đang mở trong Microsoft Excel rồi thử lại.") from exc
        finally:
            if temporary_path.exists():
                temporary_path.unlink(missing_ok=True)
    return output_path


def _write_summary(sheet, results: list[ImageResult], blur_threshold: float, reviewed: bool) -> None:
    total = len(results)
    ok = sum(1 for item in results if item.status == "OK")
    blurry = sum(1 for item in results if item.status == "BLURRY" or item.blur_score < blur_threshold)
    unreadable = sum(1 for item in results if item.status in {"UNREADABLE", "ERROR"})
    plate_count = sum(1 for item in results for plate in item.plates if _is_export_plate(plate, reviewed))
    approved = sum(1 for item in results for plate in item.plates if plate.review_approved and plate.final_text)
    corrected = sum(1 for item in results for plate in item.plates if plate.corrected_text)

    rows = [
        ("Thời gian xuất", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Tổng số ảnh", total),
        ("Ảnh đọc được biển số", ok),
        ("Tổng biển số xuất", plate_count),
        ("Ảnh mờ/cần đối chiếu", blurry),
        ("Ảnh không đọc được/lỗi", unreadable),
        ("Ngưỡng blur", blur_threshold),
    ]
    if reviewed:
        rows.extend(
            [
                ("Biển số đã Tick OK", approved),
                ("Biển số đã sửa tay", corrected),
            ]
        )
    sheet.append(["Chỉ mục", "Giá trị"])
    for row in rows:
        sheet.append(list(row))


def _write_per_image(sheet, results: list[ImageResult], media_cache: dict[tuple[str, int, int], Path], temp_dir: Path) -> None:
    max_plates = max((len([plate for plate in result.plates if plate.final_text]) for result in results), default=0)
    max_plates = max(max_plates, 3)
    plate_headers = [f"Biển số {index}" for index in range(1, max_plates + 1)]
    sheet.append(
        [
            "STT",
            "Ảnh",
            "Tên file",
            "Số biển số",
            *plate_headers,
            "Trạng thái",
            "Độ nét ảnh",
            "Kích thước",
            "Lý do/Ghi chú",
            "Đường dẫn ảnh",
        ]
    )

    for index, result in enumerate(results, start=1):
        plates = [_plate_export_text(plate) for plate in result.plates if _plate_export_text(plate)]
        padded_plates = plates + [""] * (max_plates - len(plates))
        sheet.append(
            [
                index,
                "",
                result.image_path.name,
                len(plates),
                *padded_plates,
                result.status,
                round(result.blur_score, 1),
                f"{result.width}x{result.height}" if result.width and result.height else "",
                "; ".join([part for part in [result.reason, *result.warnings] if part]),
                str(result.image_path),
            ]
        )
        _add_file_links(sheet, sheet.max_row, image_col=max_plates + 9)
        _embed_image(sheet, sheet.max_row, 2, result.image_path, media_cache, temp_dir, width=150, height=105)


def _write_readable(sheet, results: list[ImageResult], reviewed: bool, media_cache: dict[tuple[str, int, int], Path], temp_dir: Path) -> None:
    sheet.append(
        [
            "STT",
            "Ảnh",
            "Đã duyệt",
            "Tên file",
            "Đường dẫn ảnh",
            "Thư mục",
            "Biển số xuất Excel",
            "Biển số OCR",
            "Biển số OCR chuẩn hóa",
            "Độ tin cậy",
            "Độ nét ảnh",
            "Kích thước",
            "Số vùng nghi biển số",
            "Cảnh báo",
            "Crop biển số",
            "Ảnh crop",
            "OCR nguyên bản",
            "Trạng thái",
            "Lý do",
            "Loại biển đã chọn",
            "Chuỗi đã làm sạch",
            "Biển số đã định dạng",
            "Trạng thái định dạng",
            "Mẫu nhận diện",
            "Lý do cần kiểm tra",
            "Đã sửa thủ công",
        ]
    )

    row_index = 1
    for result in results:
        for plate in result.plates:
            if not _is_export_plate(plate, reviewed):
                continue
            sheet.append(
                [
                    row_index,
                    "",
                    "OK" if plate.review_approved else "",
                    result.image_path.name,
                    str(result.image_path),
                    str(result.image_path.parent),
                    _plate_export_text(plate),
                    plate.text,
                    plate.normalized_text,
                    round(plate.confidence, 1),
                    round(result.blur_score, 1),
                    f"{result.width}x{result.height}",
                    result.candidate_count,
                    "; ".join(result.warnings),
                    str(plate.crop_path) if plate.crop_path else "",
                    "",
                    plate.raw_text,
                    result.status,
                    result.reason,
                    _plate_type_value(plate.selected_plate_type),
                    plate.cleaned_text,
                    plate.formatted_text,
                    _format_status_value(plate.format_status),
                    _detected_format_value(plate.detected_format),
                    plate.format_reason,
                    plate.manual_correction,
                ]
            )
            _add_file_links(sheet, sheet.max_row, image_col=5, crop_col=15)
            _embed_image(sheet, sheet.max_row, 2, result.image_path, media_cache, temp_dir, width=120, height=86)
            if plate.crop_path:
                _embed_image(sheet, sheet.max_row, 16, plate.crop_path, media_cache, temp_dir, width=130, height=64)
            row_index += 1


def _write_special_plates(sheet, results: list[ImageResult]) -> None:
    """Write only unmatched or special/unknown plates for operator review."""

    sheet.append(
        [
            "STT",
            "Tên file",
            "Loại biển đã chọn",
            "OCR nguyên bản",
            "Chuỗi đã làm sạch",
            "Biển số đã định dạng",
            "Biển số xuất Excel",
            "Trạng thái định dạng",
            "Mẫu nhận diện",
            "Lý do cần kiểm tra",
            "Đã sửa thủ công",
            "Đường dẫn ảnh",
            "Batch ID",
        ]
    )
    row_index = 1
    for result in results:
        for plate in result.plates:
            if not _is_special_plate(plate):
                continue
            sheet.append(
                [
                    row_index,
                    result.image_path.name,
                    _plate_type_value(plate.selected_plate_type),
                    plate.raw_text or plate.text,
                    plate.cleaned_text,
                    plate.formatted_text,
                    _plate_export_text(plate),
                    _format_status_value(plate.format_status),
                    _detected_format_value(plate.detected_format),
                    plate.format_reason,
                    plate.manual_correction,
                    str(result.image_path),
                    result.batch_id,
                ]
            )
            _add_file_links(sheet, sheet.max_row, image_col=12)
            row_index += 1


def _write_review(sheet, results: list[ImageResult], blur_threshold: float, reviewed: bool, media_cache: dict[tuple[str, int, int], Path], temp_dir: Path) -> None:
    sheet.append(
        [
            "STT",
            "Ảnh",
            "Đã duyệt",
            "Tên file",
            "Đường dẫn ảnh",
            "Trạng thái",
            "Lý do",
            "Độ nét ảnh",
            "Ngưỡng blur",
            "Số vùng nghi biển số",
            "Biển số xuất Excel",
            "Biển số OCR",
            "Độ tin cậy OCR",
            "Crop đối chiếu",
            "Ảnh crop",
            "OCR nguyên bản",
            "Loại biển đã chọn",
            "Chuỗi đã làm sạch",
            "Biển số đã định dạng",
            "Trạng thái định dạng",
            "Mẫu nhận diện",
            "Lý do cần kiểm tra",
            "Đã sửa thủ công",
        ]
    )

    row_index = 1
    for result in results:
        plates = result.plates or [None]
        for plate in plates:
            needs_review = (
                result.status != "OK"
                or result.blur_score < blur_threshold
                or bool(result.warnings)
                or (reviewed and (plate is None or not plate.review_approved))
                or (plate is not None and not plate.readable and not plate.review_approved)
            )
            if not needs_review:
                continue

            sheet.append(
                [
                    row_index,
                    "",
                    "OK" if plate and plate.review_approved else "",
                    result.image_path.name,
                    str(result.image_path),
                    "ĐÃ DUYỆT" if plate and plate.review_approved else result.status,
                    "; ".join([result.reason, *(result.warnings or []), plate.reason if plate else ""]).strip("; "),
                    round(result.blur_score, 1),
                    blur_threshold,
                    result.candidate_count,
                    _plate_export_text(plate) if plate else "",
                    plate.text if plate else "",
                    round(plate.confidence, 1) if plate else "",
                    str(plate.crop_path) if plate and plate.crop_path else "",
                    "",
                    plate.raw_text if plate else result.error,
                    _plate_type_value(plate.selected_plate_type) if plate else "",
                    plate.cleaned_text if plate else "",
                    plate.formatted_text if plate else "",
                    _format_status_value(plate.format_status) if plate else "",
                    _detected_format_value(plate.detected_format) if plate else "",
                    plate.format_reason if plate else "",
                    plate.manual_correction if plate else "",
                ]
            )
            _add_file_links(sheet, sheet.max_row, image_col=5, crop_col=14)
            _embed_image(sheet, sheet.max_row, 2, result.image_path, media_cache, temp_dir, width=120, height=86)
            if plate and plate.crop_path:
                _embed_image(sheet, sheet.max_row, 15, plate.crop_path, media_cache, temp_dir, width=130, height=64)
            row_index += 1


def _write_review_all(sheet, results: list[ImageResult], media_cache: dict[tuple[str, int, int], Path], temp_dir: Path) -> None:
    sheet.append(
        [
            "STT",
            "Ảnh",
            "Đã duyệt",
            "Tên file",
            "Đường dẫn ảnh",
            "Biển số xuất Excel",
            "Biển số OCR",
            "Biển số OCR chuẩn hóa",
            "Độ tin cậy",
            "Trạng thái",
            "Lý do",
            "Crop đối chiếu",
            "Ảnh crop",
            "OCR nguyên bản",
            "Loại biển đã chọn",
            "Chuỗi đã làm sạch",
            "Biển số đã định dạng",
            "Trạng thái định dạng",
            "Mẫu nhận diện",
            "Lý do cần kiểm tra",
            "Đã sửa thủ công",
        ]
    )

    row_index = 1
    for result in results:
        for plate in result.plates:
            sheet.append(
                [
                    row_index,
                    "",
                    "OK" if plate.review_approved else "",
                    result.image_path.name,
                    str(result.image_path),
                    _plate_export_text(plate),
                    plate.text,
                    plate.normalized_text,
                    round(plate.confidence, 1),
                    "ĐÃ DUYỆT" if plate.review_approved else result.status,
                    "; ".join([part for part in [plate.reason, result.reason, *result.warnings] if part]),
                    str(plate.crop_path) if plate.crop_path else "",
                    "",
                    plate.raw_text,
                    _plate_type_value(plate.selected_plate_type),
                    plate.cleaned_text,
                    plate.formatted_text,
                    _format_status_value(plate.format_status),
                    _detected_format_value(plate.detected_format),
                    plate.format_reason,
                    plate.manual_correction,
                ]
            )
            _add_file_links(sheet, sheet.max_row, image_col=5, crop_col=12)
            _embed_image(sheet, sheet.max_row, 2, result.image_path, media_cache, temp_dir, width=120, height=86)
            if plate.crop_path:
                _embed_image(sheet, sheet.max_row, 13, plate.crop_path, media_cache, temp_dir, width=130, height=64)
            row_index += 1


def _write_all_images(sheet, results: list[ImageResult], media_cache: dict[tuple[str, int, int], Path], temp_dir: Path) -> None:
    sheet.append(
        [
            "STT",
            "Ảnh",
            "Tên file",
            "Đường dẫn ảnh",
            "Trạng thái",
            "Biển số đọc được",
            "Độ nét ảnh",
            "Kích thước",
            "Số vùng nghi biển số",
            "Lý do",
            "Cảnh báo",
        ]
    )

    for index, result in enumerate(results, start=1):
        sheet.append(
            [
                index,
                "",
                result.image_path.name,
                str(result.image_path),
                result.status,
                ", ".join(result.plate_texts),
                round(result.blur_score, 1),
                f"{result.width}x{result.height}" if result.width and result.height else "",
                result.candidate_count,
                result.reason,
                "; ".join(result.warnings),
            ]
        )
        _add_file_links(sheet, sheet.max_row, image_col=4)
        _embed_image(sheet, sheet.max_row, 2, result.image_path, media_cache, temp_dir, width=120, height=86)


def _best_plate(plates: list[PlateCandidate]) -> PlateCandidate | None:
    if not plates:
        return None
    return max(plates, key=lambda item: item.confidence)


def _plate_export_text(plate: PlateCandidate) -> str:
    return plate.export_text if plate.export_text else plate.final_text


def _plate_type_value(value: PlateType | str | None) -> str:
    return coerce_plate_type(value).value


def _format_status_value(value: PlateFormatStatus | str | None) -> str:
    try:
        return PlateFormatStatus(str(value or PlateFormatStatus.DISABLED)).value
    except ValueError:
        return PlateFormatStatus.DISABLED.value


def _detected_format_value(value: DetectedPlateFormat | str | None) -> str:
    try:
        return DetectedPlateFormat(str(value or DetectedPlateFormat.NONE)).value
    except ValueError:
        return DetectedPlateFormat.NONE.value


def _is_special_plate(plate: PlateCandidate) -> bool:
    return (
        _format_status_value(plate.format_status) == PlateFormatStatus.UNMATCHED.value
        or _detected_format_value(plate.detected_format) == DetectedPlateFormat.SPECIAL_OR_UNKNOWN.value
    )


def _is_export_plate(plate: PlateCandidate, reviewed: bool) -> bool:
    if reviewed:
        return plate.review_approved and bool(plate.final_text)
    return plate.readable and bool(plate.final_text)


def _add_file_links(sheet, row: int, image_col: int, crop_col: int | None = None) -> None:
    _set_file_link(sheet.cell(row=row, column=image_col))
    if crop_col:
        _set_file_link(sheet.cell(row=row, column=crop_col))


def _embed_image(
    sheet,
    row: int,
    column: int,
    source_path: Path,
    media_cache: dict[tuple[str, int, int], Path] | None,
    temp_dir: Path,
    width: int,
    height: int,
) -> None:
    if media_cache is None or not source_path or not source_path.exists():
        return

    thumbnail = _thumbnail(source_path, media_cache, temp_dir, width, height)
    if thumbnail is None:
        return

    image = ExcelImage(str(thumbnail))
    sheet.add_image(image, f"{get_column_letter(column)}{row}")

    current_height = sheet.row_dimensions[row].height or 15
    sheet.row_dimensions[row].height = max(current_height, height * 0.75 + 8)
    column_letter = get_column_letter(column)
    current_width = sheet.column_dimensions[column_letter].width or 10
    sheet.column_dimensions[column_letter].width = max(current_width, width / 7)


def _thumbnail(source_path: Path, media_cache: dict[tuple[str, int, int], Path], temp_dir: Path, width: int, height: int) -> Path | None:
    resolved = str(source_path.resolve())
    key = (resolved, width, height)
    if key in media_cache:
        return media_cache[key]

    try:
        with PillowImage.open(source_path) as image:
            image = ImageOps.exif_transpose(image).convert("RGB")
            image.thumbnail((width, height), PillowImage.Resampling.LANCZOS)
            digest = hashlib.sha1(f"{resolved}|{width}|{height}".encode("utf-8", errors="ignore")).hexdigest()[:16]
            thumbnail_path = temp_dir / f"{digest}.png"
            image.save(thumbnail_path, format="PNG")
    except Exception:
        return None

    media_cache[key] = thumbnail_path
    return thumbnail_path


def _set_file_link(cell) -> None:
    value = cell.value
    if not value:
        return
    path = Path(str(value))
    if path.exists():
        cell.hyperlink = path.resolve().as_uri()
        cell.style = "Hyperlink"


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        _escape_excel_formula(cell)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        status = ""
        for cell in row:
            _escape_excel_formula(cell)
            if cell.value in {"BLURRY", "UNREADABLE", "ERROR", "UNMATCHED", "SPECIAL_OR_UNKNOWN"}:
                status = str(cell.value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if status in {"BLURRY", "UNMATCHED", "SPECIAL_OR_UNKNOWN"}:
            for cell in row:
                cell.fill = WARNING_FILL
        elif status in {"UNREADABLE", "ERROR"}:
            for cell in row:
                cell.fill = ERROR_FILL

    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted = min(max(length + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted


def _escape_excel_formula(cell) -> None:
    value = cell.value
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        cell.value = "'" + value
