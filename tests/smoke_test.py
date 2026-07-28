from __future__ import annotations

import tempfile
import os
import sys
from pathlib import Path

import numpy as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from check_vehicle_ocr.config import clear_saved_api_key, load_settings, save_settings
from check_vehicle_ocr.excel_export import export_results
from check_vehicle_ocr.gemini_vision import GeminiVisionEngine, _thinking_config_for_model
from check_vehicle_ocr.image_io import collect_images, load_image
from check_vehicle_ocr.models import ImageResult, OcrAttempt, PlateCandidate
from check_vehicle_ocr.ocr import TesseractOcrEngine, find_tesseract, is_timestamp_like, looks_like_plate
from check_vehicle_ocr.paddle_ocr_engine import PaddleOcrEngine
from check_vehicle_ocr import plate_detector as plate_detector_module
from check_vehicle_ocr.processor import process_image
from check_vehicle_ocr.app import _merge_gemini_local_result, _needs_local_fallback


class FakeOcrEngine:
    def read_plate(self, crop):
        return OcrAttempt(text="30A-123.45", normalized_text="30A12345", confidence=92.0, raw_text="30A 12345", engine="fake")


class FakeGeminiEngine(GeminiVisionEngine):
    def __init__(self):
        super().__init__(api_key="fake-gemini-key", model="gemini-2.5-pro")

    def _call_model(self, image_bgr):
        return {
            "plates": [
                {
                    "plate": "70K124711",
                    "confidence": 94,
                    "vehicle": "motorbike",
                    "visibility": "clear",
                    "note": "two-line plate",
                },
                {
                    "plate": "2026-05-28 10:15",
                    "confidence": 99,
                    "vehicle": "unknown",
                    "visibility": "clear",
                    "note": "timestamp overlay",
                },
                {
                    "plate": "ABC SHOP",
                    "confidence": 95,
                    "vehicle": "unknown",
                    "visibility": "clear",
                    "note": "not a vehicle plate",
                },
            ],
            "image_blurry": False,
            "needs_review": False,
            "notes": "",
        }


class FakeRegionalEngine:
    def read_plate_regions(self, crop):
        return [
            ((20, 190, 260, 36), OcrAttempt(text="26 Thang 5, 2026", normalized_text="26THANG52026", confidence=99.0, raw_text="26 Thang 5, 2026", engine="fake")),
            ((110, 62, 250, 72), OcrAttempt(text="59-B1 581.29", normalized_text="59B158129", confidence=93.0, raw_text="59-B1 581.29", engine="fake")),
            ((430, 70, 250, 70), OcrAttempt(text="54-L1 123.45", normalized_text="54L112345", confidence=91.0, raw_text="54-L1 123.45", engine="fake")),
        ]

    def read_plate(self, crop):
        return OcrAttempt(raw_text="fallback should not be needed", engine="fake")


class CountingRegionalEngine:
    def __init__(self):
        self.batch_sizes: list[int] = []

    def read_plate_regions_batch(self, crops, **_kwargs):
        self.batch_sizes.append(len(crops))
        attempt = OcrAttempt(
            text="30A-123.45",
            normalized_text="30A12345",
            confidence=95.0,
            raw_text="30A-123.45",
            engine="fake",
            preprocess="counting_region",
        )
        return [[((290, 250, 320, 80), attempt)] for _crop in crops]

    def read_plate_regions(self, crop, **_kwargs):
        return self.read_plate_regions_batch([crop])[0]


class FakeBox:
    def __init__(self, x1, y1, x2, y2):
        self.x1 = x1
        self.y1 = y1
        self.x2 = x2
        self.y2 = y2


class FakeDetection:
    def __init__(self, x1, y1, x2, y2, confidence):
        self.bounding_box = FakeBox(x1, y1, x2, y2)
        self.confidence = confidence


class FakePlateDetector:
    def predict(self, image):
        return [
            FakeDetection(300, 250, 600, 330, 0.93),
            FakeDetection(305, 253, 598, 329, 0.80),
        ]


class FallbackGeminiEngine(GeminiVisionEngine):
    def __init__(self):
        super().__init__(api_key="fake-gemini-key", model="gemini-2.5-pro")
        self.calls: list[str] = []

    def _generate_content(self, model_name, parts, use_schema):
        self.calls.append(model_name)
        if model_name == "gemini-2.5-pro":
            raise RuntimeError("Gemini API HTTP 429: quota exceeded, limit: 0, model: gemini-2.5-pro")
        return '{"plates": [], "image_blurry": false, "needs_review": false, "notes": "fallback ok"}'


def main() -> int:
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        old_appdata = os.environ.get("APPDATA")
        os.environ["APPDATA"] = str(root / "appdata")
        try:
            save_settings(
                {"remember_key": True, "engine": "GPT Vision"},
                api_key="sk-test",
                gemini_api_key="gemini-test",
                plate_recognizer_token="plate-test",
            )
            loaded = load_settings()
            if loaded.get("api_key") != "sk-test":
                raise AssertionError("API key khong duoc luu/doc lai")
            if loaded.get("gemini_api_key") != "gemini-test":
                raise AssertionError("Gemini API key khong duoc luu/doc lai")
            if loaded.get("plate_recognizer_token") != "plate-test":
                raise AssertionError("Plate Recognizer token khong duoc luu/doc lai")
            clear_saved_api_key()
            loaded = load_settings()
            if loaded.get("api_key") or loaded.get("gemini_api_key") or loaded.get("plate_recognizer_token"):
                raise AssertionError("API key khong duoc xoa")
        finally:
            if old_appdata is None:
                os.environ.pop("APPDATA", None)
            else:
                os.environ["APPDATA"] = old_appdata

        image_path = root / "sample.jpg"
        image = Image.new("RGB", (900, 520), "white")
        draw = ImageDraw.Draw(image)
        draw.rectangle((290, 250, 610, 330), fill="white", outline="black", width=4)
        try:
            font = ImageFont.truetype("arial.ttf", 52)
        except Exception:
            font = ImageFont.load_default()
        draw.text((330, 263), "30A-123.45", fill="black", font=font)
        image.save(image_path)
        nested_dir = root / "nested"
        nested_dir.mkdir()
        nested_image_path = nested_dir / "folder_sample.jpg"
        image.save(nested_image_path)
        if collect_images(root, recursive=False) != [image_path.resolve()]:
            raise AssertionError("collect_images khong nhan Path thu muc don le")
        if collect_images([root], recursive=True) != sorted([image_path.resolve(), nested_image_path.resolve()], key=lambda item: str(item).lower()):
            raise AssertionError("collect_images khong quet dung danh sach thu muc")
        if not is_timestamp_like("26 Thang 5, 2026") or not is_timestamp_like("26 Th\u00e1ng 5, 2026"):
            raise AssertionError("Khong nhan dien duoc time mark tieng Viet")
        if not is_timestamp_like("10:27") or not is_timestamp_like("1027"):
            raise AssertionError("Khong nhan dien duoc gio tren time mark")
        if looks_like_plate("26 Thang 5, 2026"):
            raise AssertionError("Time mark bi xem nham la bien so")
        if looks_like_plate("51-X5 10:27"):
            raise AssertionError("Bien so ghep voi gio tren time mark bi xem nham la bien so")

        old_detector = plate_detector_module._DETECTOR
        old_attempted = plate_detector_module._DETECTOR_ATTEMPTED
        try:
            plate_detector_module._DETECTOR = FakePlateDetector()
            plate_detector_module._DETECTOR_ATTEMPTED = True
            onnx_candidates = plate_detector_module.detect_plate_candidates_onnx(
                image_bgr=np.zeros((520, 900, 3), dtype=np.uint8),
                max_candidates=4,
                confidence_threshold=0.25,
            )
            if len(onnx_candidates) != 1 or onnx_candidates[0].source != "onnx_plate_detector":
                raise AssertionError(f"ONNX detector adapter khong tao/dedupe dung candidate: {onnx_candidates}")
            if onnx_candidates[0].bbox[0] >= 300 or onnx_candidates[0].bbox[1] >= 250:
                raise AssertionError("ONNX detector adapter khong padding crop bien so")
        finally:
            plate_detector_module._DETECTOR = old_detector
            plate_detector_module._DETECTOR_ATTEMPTED = old_attempted

        crop_dir = root / "crops"
        result = process_image(image_path, crop_dir, FakeOcrEngine(), blur_threshold=10, confidence_threshold=45)
        if not result.plates:
            raise AssertionError("Smoke test khong tao duoc plate result")
        result.plates[0].review_approved = True
        result.plates[0].corrected_text = "=1+1"

        regional_result = process_image(image_path, root / "regional_crops", FakeRegionalEngine(), blur_threshold=10, confidence_threshold=45)
        regional_plates = sorted(plate.normalized_text for plate in regional_result.plates if plate.readable)
        if regional_plates != ["54L112345", "59B158129"]:
            raise AssertionError(f"Khong tach dung nhieu bien so hoac chua loc timestamp: {regional_plates}")

        counting_engine = CountingRegionalEngine()
        counting_result = process_image(image_path, root / "counting_crops", counting_engine, blur_threshold=10, confidence_threshold=45)
        if not counting_result.plates or sum(counting_engine.batch_sizes) != 1:
            raise AssertionError(f"Paddle primary pass da OCR lap ROI khong can thiet: {counting_engine.batch_sizes}")

        paddle_engine = PaddleOcrEngine(confidence_threshold=20)
        if paddle_engine.available:
            paddle_result = process_image(image_path, root / "paddle_crops", paddle_engine, blur_threshold=10, confidence_threshold=20)
            if not any(plate.normalized_text == "30A12345" for plate in paddle_result.plates):
                raise AssertionError("PaddleOCR smoke test khong doc duoc bien so oto")

        gemini_result = FakeGeminiEngine().analyze_image(image_path, blur_threshold=10)
        gemini_plates = [plate.normalized_text for plate in gemini_result.plates]
        if gemini_plates != ["70K124711"]:
            raise AssertionError(f"Gemini filter khong loc dung bien so: {gemini_plates}")
        if not gemini_result.plates[0].readable:
            raise AssertionError("Gemini confidence cao/clear phai duoc xem la readable")
        if _thinking_config_for_model("gemini-3-flash-preview") != {"thinkingLevel": "high"}:
            raise AssertionError("Gemini 3 thinking config khong dung")
        fallback_engine = FallbackGeminiEngine()
        image_bgr, _ = load_image(image_path)
        fallback_payload = fallback_engine._call_model(image_bgr)
        if fallback_payload.get("notes") != "fallback ok" or "gemini-2.5-flash" not in fallback_engine.calls:
            raise AssertionError(f"Gemini quota fallback khong chay dung: {fallback_engine.calls}")
        gemini_error = ImageResult(image_path=image_path, status="ERROR", reason="Gemini loi", error="429")
        local_ok = ImageResult(
            image_path=image_path,
            status="OK",
            reason="Doc duoc bien so",
            width=900,
            height=520,
            plates=[
                PlateCandidate(
                    bbox=(0, 0, 100, 40),
                    score=82,
                    source="first_pass",
                    text="30A-123.45",
                    normalized_text="30A12345",
                    confidence=82,
                    readable=True,
                )
            ],
        )
        if not _needs_local_fallback(gemini_error):
            raise AssertionError("Gemini ERROR phai kich hoat Local OCR fallback")
        merged = _merge_gemini_local_result(gemini_error, local_ok)
        if merged.status != "OK" or merged.plates[0].source != "local_ocr_fallback:first_pass":
            raise AssertionError("Merge Gemini + Local OCR fallback khong dung")

        output = root / "result.xlsx"
        export_results([result], output, blur_threshold=10, reviewed=True)
        if not output.exists():
            raise AssertionError("Smoke test khong tao duoc Excel")
        workbook = load_workbook(output)
        if "Theo_tung_anh" not in workbook.sheetnames:
            raise AssertionError("Excel thieu sheet Theo_tung_anh")
        if workbook["Bien_so_doc_duoc"]["G2"].value != "'=1+1":
            raise AssertionError("Excel chua chuyen text cong thuc thanh text an toan")

        compact_output = root / "result_without_images.xlsx"
        export_results([result], compact_output, blur_threshold=10, reviewed=True, include_images=False)
        compact_workbook = load_workbook(compact_output)
        if any(sheet._images for sheet in compact_workbook.worksheets):
            raise AssertionError("Excel tuy chon khong nhung anh van tao thumbnail")

        tesseract_path = find_tesseract()
        if tesseract_path:
            motorcycle_path = root / "motorcycle.jpg"
            motorcycle = Image.new("RGB", (900, 520), "#e5e7eb")
            draw = ImageDraw.Draw(motorcycle)
            draw.rectangle((285, 220, 625, 360), fill="white", outline="black", width=5)
            draw.text((350, 230), "70-K1", fill="black", font=font)
            draw.text((330, 292), "247.11", fill="black", font=font)
            motorcycle.save(motorcycle_path)
            tesseract_result = process_image(
                motorcycle_path,
                root / "tesseract_crops",
                TesseractOcrEngine(tesseract_path, confidence_threshold=25),
                blur_threshold=10,
                confidence_threshold=25,
            )
            if not any(plate.normalized_text == "70K124711" for plate in tesseract_result.plates):
                raise AssertionError("Tesseract smoke test khong doc duoc bien so xe may")

    print("smoke_test OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
