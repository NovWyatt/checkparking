from __future__ import annotations

import os
import queue
import sys
import tempfile
import threading
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from PIL import Image, ImageDraw


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from check_vehicle_ocr.app import CheckVehicleApp
from check_vehicle_ocr.excel_export import ExcelExportError, export_results
from check_vehicle_ocr.models import ImageResult, OcrAttempt, PlateCandidate
from check_vehicle_ocr.ocr import normalize_plate_text, plate_text_metadata
from check_vehicle_ocr.processor import process_image
from check_vehicle_ocr.services.ocr_process import OcrProcessOutcome, OcrProcessStartupError


def _image(path: Path) -> None:
    image = Image.new("RGB", (900, 520), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((290, 250, 610, 330), fill="white", outline="black", width=4)
    draw.text((330, 263), "30A-123.45", fill="black")
    image.save(path)


class RegionEngine:
    def __init__(self, attempts: list[OcrAttempt]):
        self.attempts = attempts
        self.calls: list[tuple[int, int]] = []

    def read_plate_regions_batch(self, crops, **_kwargs):
        self.calls.extend(tuple(crop.shape[:2]) for crop in crops)
        attempt = self.attempts.pop(0) if self.attempts else OcrAttempt()
        return [[((290, 250, 320, 80), attempt)] for _crop in crops]

    def read_plate_regions(self, crop, **_kwargs):
        return self.read_plate_regions_batch([crop])[0]


class ReusableProcessClient:
    init_count = 1

    def __init__(self, *, unavailable: bool = False):
        self.unavailable = unavailable
        self.start_calls = 0
        self.tasks = []

    def resume(self) -> None:
        pass

    def start(self) -> None:
        self.start_calls += 1
        if self.unavailable:
            raise OcrProcessStartupError("engine fake unavailable")

    def process(self, task, *, timeout: float = 300.0) -> OcrProcessOutcome:
        del timeout
        self.tasks.append(task)
        return OcrProcessOutcome(task.request_id, _result(task.image_path), 1.0)


def _attempt(text: str, confidence: float) -> OcrAttempt:
    return OcrAttempt(text=text, normalized_text=normalize_plate_text(text), confidence=confidence, raw_text=text)


def _result(image_path: Path) -> ImageResult:
    return ImageResult(
        image_path=image_path,
        status="OK",
        reason="ok",
        width=900,
        height=520,
        plates=[
            PlateCandidate(
                bbox=(290, 250, 320, 80),
                score=90,
                text="30A-123.45",
                corrected_text="=1+1",
                normalized_text="30A12345",
                confidence=90,
                readable=True,
                review_approved=True,
            )
        ],
    )


def test_normalization() -> None:
    for text, expected in (("30A12O45", "30A12045"), ("30A12I45", "30A12145"), ("30A12B45", "30A12845"), ("30A12S45", "30A12545")):
        cleaned, suggestions, flags, needs_review = plate_text_metadata(text)
        assert cleaned == text and expected in suggestions and flags and needs_review
        assert normalize_plate_text(text) == text
    cleaned, suggestions, flags, needs_review = plate_text_metadata("30A12345")
    assert cleaned == "30A12345" and not suggestions and not flags and not needs_review
    assert plate_text_metadata("30A\n12B45")[0] == "30A12B45"
    assert len(plate_text_metadata("30A12BSG45", max_suggestions=2)[1]) <= 2


def test_early_exit_and_engine_reuse(root: Path) -> None:
    image_path = root / "image.jpg"
    _image(image_path)
    valid = RegionEngine([_attempt("30A-123.45", 95)])
    valid_result = process_image(image_path, root / "valid", valid, blur_threshold=10, confidence_threshold=45)
    assert valid_result.plates and len(valid.calls) == 1

    missing = RegionEngine([OcrAttempt(), _attempt("30A-123.45", 95)])
    assert process_image(image_path, root / "missing", missing, blur_threshold=10, confidence_threshold=45).plates
    assert len(missing.calls) == 2

    low = RegionEngine([_attempt("not-a-plate", 15), _attempt("30A-123.45", 95)])
    assert process_image(image_path, root / "low", low, blur_threshold=10, confidence_threshold=45).plates
    assert len(low.calls) == 2 and len(set(low.calls)) == len(low.calls)

    app = object.__new__(CheckVehicleApp)
    app.event_queue = queue.Queue()
    app.stop_event = threading.Event()
    app.results = ["old-result"]
    app.worker_manager = None
    app.batch_progress = None
    app._ocr_process_client = ReusableProcessClient()
    original = CheckVehicleApp._make_engine
    try:
        CheckVehicleApp._make_engine = staticmethod(lambda *_args: (_ for _ in ()).throw(AssertionError("Paddle must stay in the child process")))
        CheckVehicleApp._worker_process(app, [image_path, image_path], root / "out.xlsx", "PaddleOCR Local", None, None, "", None, "", None, "", 10, 45, 1, "balanced")
    finally:
        CheckVehicleApp._make_engine = original
    events = list(app.event_queue.queue)
    assert app._ocr_process_client.start_calls == 1 and len(app._ocr_process_client.tasks) == 2
    assert any(event[0] == "engine_ready" for event in events)
    assert any(event[0] == "done_scan" for event in events) and app.results == ["old-result"]

    app.event_queue = queue.Queue()
    app._ocr_process_client = ReusableProcessClient(unavailable=True)
    CheckVehicleApp._worker_process(app, [image_path], root / "out.xlsx", "PaddleOCR Local", None, None, "", None, "", None, "", 10, 45, 1, "balanced")
    unavailable_events = list(app.event_queue.queue)
    assert [event[0] for event in unavailable_events] == ["ocr_tool_status", "engine_unavailable"]
    assert unavailable_events[-1][1] == "Không thể chuẩn bị công cụ nhận diện. Hãy thử mở lại ứng dụng."
    assert app.results == ["old-result"]


def test_atomic_excel(root: Path) -> None:
    image_path = root / "image.jpg"
    _image(image_path)
    result = _result(image_path)
    target = root / "result.xlsx"
    export_results([result], target, blur_threshold=10, reviewed=True, include_images=False)
    assert load_workbook(target)["Bien_so_doc_duoc"]["G2"].value == "'=1+1"
    old_bytes = target.read_bytes()

    with patch("check_vehicle_ocr.excel_export.Workbook.save", side_effect=PermissionError("locked")):
        try:
            export_results([result], target, blur_threshold=10, reviewed=True, include_images=False)
        except ExcelExportError:
            pass
        else:
            raise AssertionError("Locked Excel did not raise a friendly export error")
    assert target.read_bytes() == old_bytes
    assert not list(root.glob(".result.*.tmp.xlsx"))

    with patch("check_vehicle_ocr.excel_export._thumbnail", side_effect=AssertionError("compact mode must not thumbnail")):
        compact = root / "compact.xlsx"
        export_results([result], compact, blur_threshold=10, reviewed=True, include_images=False)
        assert all(not sheet._images for sheet in load_workbook(compact).worksheets)

    full = root / "full.xlsx"
    export_results([result], full, blur_threshold=10, reviewed=True, include_images=True)
    assert full.stat().st_size > (root / "compact.xlsx").stat().st_size
    assert load_workbook(full).sheetnames


def main() -> int:
    os.environ.setdefault("CHECK_VEHICLE_DISABLE_ONNX_DETECTOR", "1")
    with tempfile.TemporaryDirectory(prefix="check_vehicle_stability_") as temporary:
        root = Path(temporary)
        test_normalization()
        test_early_exit_and_engine_reuse(root)
        test_atomic_excel(root)
    print("performance_stability_test OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
