from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from check_vehicle_ocr.config import migrate_settings
from check_vehicle_ocr.excel_export import export_results
from check_vehicle_ocr.models import BatchSession, ImageResult, PlateCandidate
from check_vehicle_ocr.plate_formatting import (
    DetectedPlateFormat,
    PlateFormatStatus,
    PlateType,
    clean_plate_for_formatting,
    format_plate,
    reformat_manual_correction,
)


def _assert_format(text: str, plate_type: PlateType, expected: str, detected: DetectedPlateFormat) -> None:
    result = format_plate(text, plate_type)
    assert result.raw_text == text
    assert result.formatted_text == expected
    assert result.export_text == expected
    assert result.format_status is PlateFormatStatus.FORMATTED
    assert result.detected_format is detected
    assert result.needs_review is False


def _assert_unmatched(text: str, plate_type: PlateType) -> None:
    result = format_plate(text, plate_type)
    assert result.raw_text == text
    assert result.formatted_text == ""
    assert result.export_text == text
    assert result.format_status is PlateFormatStatus.UNMATCHED
    assert result.detected_format is DetectedPlateFormat.SPECIAL_OR_UNKNOWN
    assert result.needs_review is True


def main() -> int:
    assert clean_plate_for_formatting("59X1-123.45\n") == "59X112345"
    assert clean_plate_for_formatting("59 MN 12345") == "59MN12345"
    assert clean_plate_for_formatting("O1-BS") == "O1BS"

    _assert_format("59X112345", PlateType.MOTORCYCLE, "59X1-12345", DetectedPlateFormat.MOTORCYCLE_LETTER_DIGIT)
    _assert_format("59X11234", PlateType.MOTORCYCLE, "59X1-1234", DetectedPlateFormat.MOTORCYCLE_LETTER_DIGIT)
    _assert_format("59MN12345", PlateType.MOTORCYCLE, "59MN-12345", DetectedPlateFormat.MOTORCYCLE_TWO_LETTERS)
    _assert_format("59MN1234", PlateType.MOTORCYCLE, "59MN-1234", DetectedPlateFormat.MOTORCYCLE_TWO_LETTERS)
    _assert_format("59.MN-12345", PlateType.MOTORCYCLE, "59MN-12345", DetectedPlateFormat.MOTORCYCLE_TWO_LETTERS)
    _assert_format("59X1-12345", PlateType.MOTORCYCLE, "59X1-12345", DetectedPlateFormat.MOTORCYCLE_LETTER_DIGIT)
    _assert_format("59X12345", PlateType.CAR, "59X-12345", DetectedPlateFormat.CAR_STANDARD)
    _assert_format("59X1234", PlateType.CAR, "59X-1234", DetectedPlateFormat.CAR_STANDARD)

    _assert_unmatched("49MD112345", PlateType.MOTORCYCLE)
    _assert_unmatched("59-110-MN-123", PlateType.MOTORCYCLE)
    for invalid in ("", "ABC", "123456", "59X1", "59X11234567", "59@X#12345"):
        format_plate(invalid, PlateType.CAR)

    disabled = format_plate("59X112345", PlateType.NONE)
    assert disabled.format_status is PlateFormatStatus.DISABLED
    assert disabled.export_text == "59X112345"

    manual_valid = reformat_manual_correction("59X112345", PlateType.MOTORCYCLE)
    assert manual_valid.format_status is PlateFormatStatus.MANUAL
    assert manual_valid.export_text == "59X1-12345"
    assert manual_valid.manual_correction == "59X112345"
    manual_special = reformat_manual_correction("49MD1-12345", PlateType.MOTORCYCLE)
    assert manual_special.format_status is PlateFormatStatus.UNMATCHED
    assert manual_special.export_text == "49MD1-12345"
    assert manual_special.needs_review is True

    migrated = migrate_settings({})
    assert migrated["last_plate_type"] == PlateType.NONE.value
    assert migrate_settings({"last_plate_type": "unexpected"})["last_plate_type"] == PlateType.NONE.value

    batch = BatchSession("batch-format-test", PlateType.MOTORCYCLE, "2026-07-29T10:00:00+07:00", 2)
    standard = PlateCandidate(
        bbox=(0, 0, 1, 1),
        score=90,
        text="59X1-12345",
        raw_text="59X1-12345",
        readable=True,
    )
    standard.apply_plate_formatting(batch.selected_plate_type)
    assert standard.raw_text == "59X1-12345"
    assert standard.cleaned_text == "59X112345"
    assert standard.formatted_text == "59X1-12345"
    assert standard.export_text == "59X1-12345"
    assert standard.selected_plate_type is PlateType.MOTORCYCLE

    special = PlateCandidate(
        bbox=(0, 0, 1, 1),
        score=80,
        text="49MD-112.345",
        raw_text="49MD112345",
        readable=True,
    )
    special.apply_plate_formatting(batch.selected_plate_type)
    assert special.raw_text == "49MD112345"
    assert special.export_text == "49MD112345"
    assert special.needs_review is True

    # Reformatting a candidate performs no OCR: only the stored text changes.
    standard.apply_plate_formatting(PlateType.CAR)
    assert standard.export_text == "59X1-12345"
    assert standard.format_status is PlateFormatStatus.UNMATCHED
    assert standard.needs_review is True

    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        image_path = root / "plate.jpg"
        Image.new("RGB", (16, 8), "white").save(image_path)
        result = ImageResult(
            image_path=image_path,
            status="OK",
            reason="Đọc được biển số",
            width=16,
            height=8,
            plates=[standard, special],
            batch_id=batch.batch_id,
            selected_plate_type=batch.selected_plate_type,
            batch_started_at=batch.started_at,
            batch_total_images=batch.total_images,
        )
        # Restore one standard result for the main export assertion.
        standard.apply_plate_formatting(PlateType.MOTORCYCLE)
        assert standard.needs_review is False
        output = root / "plates.xlsx"
        export_results([result], output, blur_threshold=10, include_images=False)
        workbook = load_workbook(output)
        readable = workbook["Bien_so_doc_duoc"]
        headers = [cell.value for cell in readable[1]]
        for header in (
            "Loại biển đã chọn",
            "OCR nguyên bản",
            "Chuỗi đã làm sạch",
            "Biển số đã định dạng",
            "Biển số xuất Excel",
            "Trạng thái định dạng",
            "Mẫu nhận diện",
            "Lý do cần kiểm tra",
            "Đã sửa thủ công",
        ):
            assert header in headers
        assert readable["G2"].value == "59X1-12345"
        special_sheet = workbook["Bien_so_dac_biet"]
        assert special_sheet.max_row == 2
        assert special_sheet["G2"].value == "49MD112345"
        assert special_sheet["H2"].value == PlateFormatStatus.UNMATCHED.value

    print("plate_formatting_test OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
