from __future__ import annotations

import tempfile
import sys
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from check_vehicle_ocr.hybrid_review import AiReviewPolicy, should_send_to_ai
from check_vehicle_ocr.excel_export import export_results
from check_vehicle_ocr.models import ExpectedPlateCount, OcrAttempt, PlateCandidate
from check_vehicle_ocr.paddle_ocr_engine import _can_group_plate_lines
from check_vehicle_ocr.plate_formatting import PlateFormatStatus, PlateType
from check_vehicle_ocr.plate_selection import is_plate_like_candidate
from check_vehicle_ocr import processor


def _overlay_fixture(path: Path) -> np.ndarray:
    """Create a no-network regression image resembling a phone-camera overlay."""

    frame = np.full((540, 960, 3), 215, dtype=np.uint8)
    cv2.rectangle(frame, (160, 145), (800, 470), (80, 80, 80), thickness=-1)
    cv2.rectangle(frame, (315, 300), (645, 382), (245, 245, 245), thickness=-1)
    cv2.rectangle(frame, (315, 300), (645, 382), (15, 15, 15), thickness=3)
    cv2.putText(frame, "59X1-12345", (345, 355), cv2.FONT_HERSHEY_SIMPLEX, 1.1, (0, 0, 0), 3, cv2.LINE_AA)
    cv2.rectangle(frame, (0, 0), (960, 42), (18, 18, 18), thickness=-1)
    cv2.rectangle(frame, (0, 495), (960, 540), (18, 18, 18), thickness=-1)
    cv2.putText(frame, "2026-07-29 10:27:03", (14, 29), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1, cv2.LINE_AA)
    cv2.putText(frame, "Diem danh - 123 Duong Mau - watermark LOL", (14, 525), cv2.FONT_HERSHEY_SIMPLEX, 0.46, (255, 255, 255), 1, cv2.LINE_AA)
    assert cv2.imwrite(str(path), frame)
    return frame


class _Detector:
    def __init__(self, boxes: list[tuple[int, int, int, int, float]]) -> None:
        self.boxes = boxes
        self.calls = 0

    def __call__(self, _image, *, max_candidates: int, confidence_threshold: float):
        self.calls += 1
        return [
            PlateCandidate(
                bbox=(x, y, width, height),
                score=confidence,
                detector_confidence=confidence,
                source="onnx_plate_detector",
            )
            for x, y, width, height, confidence in self.boxes[:max_candidates]
            if confidence >= confidence_threshold
        ]


class _OverlayRegionEngine:
    def __init__(self) -> None:
        self.crop_shapes: list[tuple[int, int]] = []
        self.full_scene_calls = 0

    def read_plate_regions(self, crop: np.ndarray):
        height, width = crop.shape[:2]
        self.crop_shapes.append((width, height))
        if width >= 900:
            self.full_scene_calls += 1
        return [
            ((20, 15, 245, 48), OcrAttempt(raw_text="59X112345", text="59X112345", normalized_text="59X112345", confidence=96.0, engine="fake")),
            ((2, 2, 20, 12), OcrAttempt(raw_text="M", text="M", normalized_text="M", confidence=99.0, engine="fake")),
            ((2, 20, 30, 12), OcrAttempt(raw_text="LOL", text="LOL", normalized_text="LOL", confidence=99.0, engine="fake")),
            ((2, 35, 130, 12), OcrAttempt(raw_text="2026-07-29 10:27:03", text="2026-07-29 10:27:03", normalized_text="20260729102703", confidence=99.0, engine="fake")),
            ((2, 50, 140, 12), OcrAttempt(raw_text="Diem danh 123 Duong Mau", text="Diem danh 123 Duong Mau", normalized_text="DIEMDANH123DUONGMAU", confidence=98.0, engine="fake")),
            ((2, 65, 90, 12), OcrAttempt(raw_text="C-F453-BE", text="C-F453-BE", normalized_text="CF453BE", confidence=97.0, engine="fake")),
        ]


class _TwoPlateEngine:
    def __init__(self) -> None:
        self.calls = 0

    def read_plate_regions(self, _crop: np.ndarray):
        self.calls += 1
        raw = "59X112345" if self.calls == 1 else "51A12345"
        return [((10, 10, 220, 52), OcrAttempt(raw_text=raw, text=raw, normalized_text=raw, confidence=95.0, engine="fake"))]


class _CropOnlyTesseract:
    def __init__(self) -> None:
        self.received_shapes: list[tuple[int, int]] = []

    def read_plate(self, crop: np.ndarray):
        self.received_shapes.append((crop.shape[1], crop.shape[0]))
        return OcrAttempt(raw_text="59X112345", text="59X112345", normalized_text="59X112345", confidence=95.0, engine="tesseract")


def _with_detector(detector: _Detector):
    previous = processor.detect_plate_candidates_onnx
    processor.detect_plate_candidates_onnx = detector
    return previous


def _assert_one_primary_with_overlays(root: Path, frame: np.ndarray, image_path: Path):
    detector = _Detector([(300, 285, 370, 115, 95.0)])
    previous = _with_detector(detector)
    try:
        engine = _OverlayRegionEngine()
        result = processor.process_image(
            image_path,
            root / "crops",
            engine,
            blur_threshold=10,
            confidence_threshold=70,
            paddle_scan_mode="fast",
            image_bgr=frame,
            image_size=(frame.shape[1], frame.shape[0]),
            selected_plate_type=PlateType.MOTORCYCLE,
            expected_plate_count=ExpectedPlateCount.ONE,
        )
    finally:
        processor.detect_plate_candidates_onnx = previous

    assert result.status == "OK"
    assert len(result.plates) == 1
    assert result.primary_plate is result.plates[0]
    assert result.plates[0].export_text == "59X1-12345"
    assert result.plates[0].format_status is PlateFormatStatus.FORMATTED
    assert not result.plates[0].needs_review
    assert result.pipeline_metrics["detector_calls"] == 1
    assert result.pipeline_metrics["crop_ocr_calls"] == 1
    assert result.pipeline_metrics["full_scene_ocr_calls"] == 0
    assert result.pipeline_metrics["tesseract_calls"] == 0
    assert result.pipeline_metrics["ai_calls"] == 0
    assert result.pipeline_metrics["candidates_before_filter"] == 6
    assert result.pipeline_metrics["candidates_after_filter"] == 1
    assert engine.full_scene_calls == 0
    assert {candidate.raw_text for candidate in result.rejected_candidates} >= {
        "M",
        "LOL",
        "2026-07-29 10:27:03",
        "Diem danh 123 Duong Mau",
        "C-F453-BE",
    }
    assert all(candidate.format_status is PlateFormatStatus.REJECTED_NOISE for candidate in result.rejected_candidates)
    send_to_ai, _reason = should_send_to_ai(result, AiReviewPolicy.NEEDS_REVIEW, confidence_threshold=70)
    assert not send_to_ai
    output = root / "overlay-result.xlsx"
    export_results([result], output, blur_threshold=70, include_images=False)
    workbook = load_workbook(output)
    assert workbook["Bien_so_dac_biet"].max_row == 1
    assert workbook["Bien_so_doc_duoc"].max_row == 2
    return result


def _assert_multi_plate_policy(root: Path, frame: np.ndarray, image_path: Path) -> None:
    detector = _Detector([(170, 230, 300, 80, 95.0), (500, 300, 300, 80, 94.0)])
    previous = _with_detector(detector)
    try:
        one = processor.process_image(
            image_path,
            root / "one",
            _TwoPlateEngine(),
            blur_threshold=10,
            confidence_threshold=70,
            paddle_scan_mode="balanced",
            image_bgr=frame,
            image_size=(frame.shape[1], frame.shape[0]),
            selected_plate_type=PlateType.MOTORCYCLE,
            expected_plate_count=ExpectedPlateCount.ONE,
        )
        many = processor.process_image(
            image_path,
            root / "many",
            _TwoPlateEngine(),
            blur_threshold=10,
            confidence_threshold=70,
            paddle_scan_mode="balanced",
            image_bgr=frame,
            image_size=(frame.shape[1], frame.shape[0]),
            selected_plate_type=PlateType.MOTORCYCLE,
            expected_plate_count=ExpectedPlateCount.MULTIPLE,
        )
    finally:
        processor.detect_plate_candidates_onnx = previous
    assert len(one.plates) == 1
    assert len(many.plates) == 2
    assert many.pipeline_metrics["full_scene_ocr_calls"] == 0


def _assert_tesseract_never_sees_scene(root: Path, frame: np.ndarray, image_path: Path) -> None:
    detector = _Detector([(300, 285, 370, 115, 95.0)])
    previous = _with_detector(detector)
    try:
        engine = _CropOnlyTesseract()
        result = processor.process_image(
            image_path,
            root / "tesseract",
            engine,
            blur_threshold=10,
            confidence_threshold=70,
            paddle_scan_mode="balanced",
            image_bgr=frame,
            image_size=(frame.shape[1], frame.shape[0]),
            selected_plate_type=PlateType.MOTORCYCLE,
        )
    finally:
        processor.detect_plate_candidates_onnx = previous
    assert result.plates
    assert result.pipeline_metrics["full_scene_ocr_calls"] == 0
    assert result.pipeline_metrics["tesseract_calls"] == 1
    assert engine.received_shapes and all(width < frame.shape[1] for width, _height in engine.received_shapes)


def main() -> int:
    for noise in ("M", "Y", "I", "LOL", "TEE", "C-F453-BE", "2026-07-29 10:27:03", "Diem danh 123 Duong Mau"):
        assert not is_plate_like_candidate(noise), noise
    for special in ("49MD112345", "59-110-MN-123"):
        assert is_plate_like_candidate(special), special
    assert _can_group_plate_lines((20, 20, 180, 35), (25, 62, 170, 35))
    assert not _can_group_plate_lines((20, 20, 180, 35), (600, 250, 170, 35))

    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        image_path = root / "overlay.jpg"
        frame = _overlay_fixture(image_path)
        _assert_one_primary_with_overlays(root, frame, image_path)
        _assert_multi_plate_policy(root, frame, image_path)
        _assert_tesseract_never_sees_scene(root, frame, image_path)

        detector = _Detector([(300, 285, 370, 115, 95.0)])
        previous = _with_detector(detector)
        try:
            results = []
            for index in range(18):
                path = root / f"overlay_{index:02d}.jpg"
                assert cv2.imwrite(str(path), frame)
                results.append(
                    processor.process_image(
                        path,
                        root / "batch",
                        _OverlayRegionEngine(),
                        blur_threshold=10,
                        confidence_threshold=70,
                        paddle_scan_mode="fast",
                        image_bgr=frame,
                        image_size=(frame.shape[1], frame.shape[0]),
                        selected_plate_type=PlateType.MOTORCYCLE,
                        expected_plate_count=ExpectedPlateCount.ONE,
                    )
                )
        finally:
            processor.detect_plate_candidates_onnx = previous
        assert len(results) == 18
        assert sum(len(result.plates) for result in results) == 18
        assert all(not result.primary_plate.needs_review for result in results)

    print("false_positive_regression_test OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
