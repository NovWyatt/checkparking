from __future__ import annotations

import sys
import tempfile
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from check_vehicle_ocr.candidate_scoring import apply_fallback_selection, choose_engine_candidate
from check_vehicle_ocr.excel_export import export_results
from check_vehicle_ocr.models import ImageResult, PlateCandidate
from check_vehicle_ocr.plate_formatting import PlateType


def _plate(text: str, confidence: float, *, engine: str) -> PlateCandidate:
    return PlateCandidate(
        bbox=(0, 0, 100, 30),
        score=confidence,
        text=text,
        normalized_text=text.replace("-", ""),
        raw_text=text,
        confidence=confidence,
        readable=True,
        selected_engine=engine,
    )


def test_candidate_decisions() -> None:
    agree = choose_engine_candidate("59X112345", 82, "59X112345", 76, PlateType.MOTORCYCLE)
    assert agree.engine == "paddle+tesseract" and not agree.needs_review
    tess = choose_engine_candidate("59X112345", 35, "59MN12345", 92, PlateType.MOTORCYCLE)
    assert tess.engine == "tesseract" and not tess.needs_review
    uncertain = choose_engine_candidate("59X112345", 79, "59MN12345", 78, PlateType.MOTORCYCLE)
    assert uncertain.needs_review


def test_fallback_data_and_excel_sheet() -> None:
    paddle = ImageResult(Path("paddle.jpg"), "OK", "", plates=[_plate("59X112345", 42, engine="paddleocr")])
    tess = ImageResult(Path("paddle.jpg"), "OK", "", plates=[_plate("59MN12345", 92, engine="tesseract")])
    merged = apply_fallback_selection(paddle, tess, PlateType.MOTORCYCLE)
    plate = merged.plates[0]
    assert plate.paddle_raw == "59X112345" and plate.tesseract_raw == "59MN12345"
    assert plate.selected_engine == "tesseract" and plate.text == "59MN12345"
    with tempfile.TemporaryDirectory() as temporary:
        output = Path(temporary) / "result.xlsx"
        export_results([merged], output, blur_threshold=20, include_images=False)
        from openpyxl import load_workbook

        workbook = load_workbook(output)
        assert "So_sanh_OCR" in workbook.sheetnames
        values = [cell.value for cell in workbook["So_sanh_OCR"][2]]
        assert "paddleocr" not in values and "tesseract" in values


def main() -> int:
    test_candidate_decisions()
    test_fallback_data_and_excel_sheet()
    print("v190_candidate_scoring_test OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
