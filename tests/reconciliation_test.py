from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from check_vehicle_ocr.reconciliation import (
    export_reconciliation_report,
    reconcile_workbooks,
    write_reconciliation_template,
)


def _write_source(path: Path, header: str, plates: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh_sach"
    sheet.append([header])
    for plate in plates:
        sheet.append([plate])
    workbook.save(path)
    workbook.close()


def main() -> int:
    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        template = root / "mau_bao_phi.xlsx"
        write_reconciliation_template(template, "Báo phí")
        template_book = load_workbook(template)
        if template_book.sheetnames != ["Danh_sach", "Huong_dan"]:
            raise AssertionError("File mẫu phải có sheet dán dữ liệu và hướng dẫn")
        if template_book["Danh_sach"]["A1"].value != "Biển số":
            raise AssertionError("File mẫu không có cột Biển số rõ ràng")
        template_book.close()

        ocr = root / "ocr.xlsx"
        fee = root / "bao_phi.xlsx"
        software = root / "phan_mem.xlsx"
        _write_source(
            ocr,
            "Biển số xuất Excel",
            [
                "30A-123.45",
                "51F-123.45",
                "59X1-999.99",
                "43A-111.11",
                "30A-555.55",
                "50A-222.22",
                "30A-765.43",
                "30A-987.65",
            ],
        )
        _write_source(
            fee,
            "Biển số",
            [
                "30A12345",
                "30A12345",
                "51F-123.46",
                "30A-555.54",
                "30A-555.56",
                "30A-765.3",
                "60A-000.01",
                "=SUM(A1:A2)",
            ],
        )
        _write_source(software, "Biển số", ["59X1-999.99", "50A-222.23", "30A-9876.65", "70A-000.01"])

        report = reconcile_workbooks(ocr, fee, software, suffix_length=4)
        if len(report.rows) != 8:
            raise AssertionError(f"Phải đọc 8 biển OCR, nhận được {len(report.rows)}")
        statuses = [(row.fee.status, row.software.status, row.conclusion) for row in report.rows]
        expected = [
            ("exact", "skipped", "Khớp báo phí"),
            ("near", "skipped", "Khớp gần báo phí"),
            ("missing", "exact", "Có phần mềm, không báo phí"),
            ("missing", "missing", "Không có trên cả hai nguồn"),
            ("ambiguous", "missing", "Cần xác nhận báo phí"),
            ("missing", "near", "Có gần đúng phần mềm, không báo phí"),
            ("near", "skipped", "Khớp gần báo phí"),
            ("missing", "near", "Có gần đúng phần mềm, không báo phí"),
        ]
        if statuses != expected:
            raise AssertionError(f"Phân loại đối chiếu sai: {statuses}")
        if "Thiếu hoặc dư 1 ký tự" not in report.rows[6].fee.note:
            raise AssertionError("Thiếu một ký tự ở báo phí phải được nhận diện là khớp gần")
        if "Thiếu hoặc dư 1 ký tự" not in report.rows[7].software.note:
            raise AssertionError("Dư một ký tự ở phần mềm phải được nhận diện là khớp gần")
        if not any(source == "Báo phí" and normalized == "30A12345" for source, normalized, _records in report.duplicates):
            raise AssertionError("Phải báo biển trùng trong nguồn báo phí")

        output = root / "doi_chieu.xlsx"
        export_reconciliation_report(report, output)
        workbook = load_workbook(output)
        expected_sheets = {
            "Tổng_quan",
            "Kết_quả_chính",
            "Khớp_báo_phí",
            "Khớp_gần_báo_phí",
            "Phần_mềm_không_báo_phí",
            "Không_có_cả_hai",
            "Cần_xác_nhận",
            "Trùng_lặp",
            "Dư_báo_phí",
            "Dư_phần_mềm",
        }
        if not expected_sheets.issubset(workbook.sheetnames):
            raise AssertionError(f"Thiếu sheet báo cáo: {workbook.sheetnames}")
        if workbook["Kết_quả_chính"]["M2"].value != "Khớp báo phí":
            raise AssertionError("Sheet kết quả chính không giữ kết luận đối chiếu")
        if workbook["Khớp_gần_báo_phí"].max_row != 3:
            raise AssertionError("Các khớp gần báo phí phải được tách thành sheet riêng")
        workbook.close()

        fee_only = reconcile_workbooks(ocr, fee, suffix_length=3)
        if fee_only.compare_software:
            raise AssertionError("Đối chiếu chỉ báo phí không được yêu cầu file phần mềm")
        fee_only_output = root / "doi_chieu_bao_phi.xlsx"
        export_reconciliation_report(fee_only, fee_only_output)
        fee_only_book = load_workbook(fee_only_output)
        if "Phần_mềm_không_báo_phí" in fee_only_book.sheetnames:
            raise AssertionError("Báo cáo chỉ báo phí không được tạo sheet phần mềm")
        fee_only_book.close()

    print("reconciliation_test OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
